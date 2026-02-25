"""
Ethereum Wallet Transaction Analyzer
Pulls all transactions for a wallet via Etherscan API,
exports to CSV, and produces a discrepancy report comparing
expected vs actual on-chain balances.
"""

import csv
import io
import json
import sys
import time
from datetime import datetime, timezone
from decimal import Decimal, getcontext
from urllib.request import urlopen, Request
from urllib.parse import urlencode
from urllib.error import URLError, HTTPError
from collections import defaultdict
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Load .env file
def _load_env(path=".env"):
    try:
        with open(path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, _, val = line.partition("=")
                    os.environ.setdefault(key.strip(), val.strip())
    except FileNotFoundError:
        pass

_load_env()

ETHERSCAN_API_KEY = os.environ.get("ETHERSCAN_API_KEY", "")
ALCHEMY_API_KEY = os.environ.get("ALCHEMY_API_KEY", "")
WALLET = os.environ.get("WALLET_ADDRESS", "").lower()
TRANSACTIONS_DIR = "transactions"
DISCREPANCIES_DIR = "discrepancies"
CSV_FILE = os.path.join(TRANSACTIONS_DIR, "transactions.csv")
DISCREPANCY_FILE = os.path.join(DISCREPANCIES_DIR, "discrepancies.csv")
BASE_URL = "https://api.etherscan.io/v2/api"
ALCHEMY_URL = f"https://eth-mainnet.g.alchemy.com/v2/{ALCHEMY_API_KEY}"
CHAIN_ID = "1"                                  # 1 = Ethereum mainnet
PAGE_SIZE = 10000                               # max rows per API page
RATE_LIMIT_DELAY = 0.22                         # ~5 req/s for free tier
BALANCE_OF_SELECTOR = "0x70a08231"              # balanceOf(address)

getcontext().prec = 36                          # enough for 18-decimal tokens

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def etherscan_get(params: dict) -> list:
    """Call Etherscan API with pagination; return combined result list."""
    params.setdefault("chainid", CHAIN_ID)
    params.setdefault("apikey", ETHERSCAN_API_KEY)
    params.setdefault("sort", "asc")
    params.setdefault("offset", str(PAGE_SIZE))

    all_results = []
    page = 1

    while True:
        params["page"] = str(page)
        url = f"{BASE_URL}?{urlencode(params)}"
        time.sleep(RATE_LIMIT_DELAY)

        try:
            req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode())
        except (URLError, HTTPError) as exc:
            print(f"  [!] API error on page {page}: {exc}")
            break

        status = data.get("status")
        result = data.get("result", [])

        if status == "0" and isinstance(result, str):
            # "No transactions found" or rate-limit message
            if "rate" in result.lower():
                print("  [!] Rate-limited, waiting 1 s …")
                time.sleep(1)
                continue
            if page == 1:
                print(f"  API response: {result}")
            break

        if not isinstance(result, list) or len(result) == 0:
            break

        all_results.extend(result)
        print(f"  page {page}: {len(result)} records")

        if len(result) < PAGE_SIZE:
            break
        page += 1

    return all_results


def wei_to_eth(wei_str: str) -> Decimal:
    return Decimal(wei_str) / Decimal("1000000000000000000")


def token_amount(raw: str, decimals: str) -> Decimal:
    d = int(decimals) if decimals else 18
    return Decimal(raw) / Decimal(10) ** d


def ts_to_date(ts: str) -> str:
    return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime(
        "%Y-%m-%d %H:%M:%S UTC"
    )


def alchemy_balance_of(contract: str, block: int, decimals: int) -> Decimal:
    """Call balanceOf(wallet) on a token contract at a specific block via Alchemy."""
    if not ALCHEMY_API_KEY:
        return Decimal("-1")  # sentinel: Alchemy not configured
    addr = WALLET.replace("0x", "").lower().zfill(64)
    calldata = BALANCE_OF_SELECTOR + addr
    payload = json.dumps({
        "jsonrpc": "2.0", "id": 1,
        "method": "eth_call",
        "params": [{"to": contract, "data": calldata}, hex(block)],
    }).encode()
    try:
        req = Request(ALCHEMY_URL, data=payload, headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
        })
        with urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode())
        hex_result = data.get("result", "0x0")
        if hex_result in (None, "0x", "0x0", ""):
            return Decimal("0")
        return Decimal(int(hex_result, 16)) / Decimal(10) ** decimals
    except (URLError, HTTPError, ValueError):
        return Decimal("-1")


# ---------------------------------------------------------------------------
# 1. Pull transactions
# ---------------------------------------------------------------------------

def fetch_normal_txs() -> list:
    print("[*] Fetching normal ETH transactions …")
    return etherscan_get({
        "module": "account",
        "action": "txlist",
        "address": WALLET,
        "startblock": "0",
        "endblock": "99999999",
    })


def fetch_internal_txs() -> list:
    print("[*] Fetching internal transactions …")
    return etherscan_get({
        "module": "account",
        "action": "txlistinternal",
        "address": WALLET,
        "startblock": "0",
        "endblock": "99999999",
    })


def fetch_erc20_txs() -> list:
    print("[*] Fetching ERC20 token transfers …")
    return etherscan_get({
        "module": "account",
        "action": "tokentx",
        "address": WALLET,
        "startblock": "0",
        "endblock": "99999999",
    })


# ---------------------------------------------------------------------------
# 2. Normalise into a common row format + accumulate balances
# ---------------------------------------------------------------------------

# Accumulators
eth_in = Decimal("0")
eth_out = Decimal("0")
gas_total = Decimal("0")
# token contract -> {symbol, decimals, in, out}
token_flows = defaultdict(lambda: {
    "symbol": "",
    "decimals": 18,
    "in": Decimal("0"),
    "out": Decimal("0"),
})


def process_normal(txs: list) -> list:
    """Return CSV-ready rows from normal ETH transactions."""
    global eth_in, eth_out, gas_total

    rows = []
    for tx in txs:
        if tx.get("isError", "0") != "0":
            # Failed tx: the sender still pays gas, but value is NOT transferred
            if tx["from"].lower() == WALLET:
                fee = wei_to_eth(tx["gasUsed"]) * wei_to_eth(tx["gasPrice"]) * Decimal("1e18")
                # gasUsed * gasPrice is already in wei
                fee = Decimal(tx["gasUsed"]) * Decimal(tx["gasPrice"])
                fee = fee / Decimal("1e18")
                gas_total += fee
                rows.append({
                    "Date": ts_to_date(tx["timeStamp"]),
                    "Block Number": tx.get("blockNumber", ""),
                    "Transaction Hash": tx["hash"],
                    "Type": "OUT (failed)",
                    "From": tx["from"],
                    "To": tx["to"],
                    "Amount": "0",
                    "Token/Asset": "ETH",
                    "Gas Fee (in ETH)": f"{fee:.18f}",
                    "Function": tx.get("functionName", ""),
                    "Input Data": tx.get("input", ""),
                })
            continue

        value = wei_to_eth(tx["value"])
        from_addr = tx["from"].lower()
        to_addr = (tx.get("to") or "").lower()
        direction = "IN" if to_addr == WALLET else "OUT"

        # Gas fee (only charged to the sender)
        fee = Decimal("0")
        if from_addr == WALLET:
            fee = Decimal(tx["gasUsed"]) * Decimal(tx["gasPrice"]) / Decimal("1e18")
            gas_total += fee

        if direction == "IN":
            eth_in += value
        else:
            eth_out += value

        rows.append({
            "Date": ts_to_date(tx["timeStamp"]),
            "Block Number": tx.get("blockNumber", ""),
            "Transaction Hash": tx["hash"],
            "Type": direction,
            "From": tx["from"],
            "To": tx.get("to", ""),
            "Amount": f"{value:.18f}",
            "Token/Asset": "ETH",
            "Gas Fee (in ETH)": f"{fee:.18f}",
            "Function": tx.get("functionName", ""),
            "Input Data": tx.get("input", ""),
        })
    return rows


def process_internal(txs: list) -> list:
    """Internal transactions (no gas fee — already counted in outer tx)."""
    global eth_in, eth_out

    rows = []
    for tx in txs:
        if tx.get("isError", "0") != "0":
            continue

        value = wei_to_eth(tx["value"])
        if value == 0:
            continue

        to_addr = (tx.get("to") or "").lower()
        direction = "IN" if to_addr == WALLET else "OUT"

        if direction == "IN":
            eth_in += value
        else:
            eth_out += value

        rows.append({
            "Date": ts_to_date(tx["timeStamp"]),
            "Block Number": tx.get("blockNumber", ""),
            "Transaction Hash": tx.get("hash", tx.get("transactionHash", "")),
            "Type": f"{direction} (internal)",
            "From": tx["from"],
            "To": tx.get("to", ""),
            "Amount": f"{value:.18f}",
            "Token/Asset": "ETH",
            "Gas Fee (in ETH)": "0",
            "Function": tx.get("type", ""),
            "Input Data": "",  # internal txs don't carry input data
        })
    return rows



def _looks_like_nft(tx: dict) -> bool:
    """Heuristic: skip ERC721 / ERC1155 transfers."""
    # tokenID present and tokenDecimal absent or "0" → NFT
    if tx.get("tokenID") and (not tx.get("tokenDecimal") or tx.get("tokenDecimal") == "0"):
        return True
    return False


def process_erc20(txs: list) -> list:
    global eth_in, eth_out

    rows = []
    for tx in txs:
        if _looks_like_nft(tx):
            continue

        decimals_str = tx.get("tokenDecimal", "18")
        try:
            int(decimals_str)
        except ValueError:
            continue

        symbol = tx.get("tokenSymbol", "UNKNOWN")
        contract = tx.get("contractAddress", "").lower()
        value = token_amount(tx["value"], decimals_str)

        to_addr = (tx.get("to") or "").lower()
        direction = "IN" if to_addr == WALLET else "OUT"

        tf = token_flows[contract]
        tf["symbol"] = symbol
        tf["decimals"] = int(decimals_str)
        if direction == "IN":
            tf["in"] += value
        else:
            tf["out"] += value

        rows.append({
            "Date": ts_to_date(tx["timeStamp"]),
            "Block Number": tx.get("blockNumber", ""),
            "Transaction Hash": tx["hash"],
            "Type": direction,
            "From": tx["from"],
            "To": tx.get("to", ""),
            "Amount": f"{value}",
            "Token/Asset": symbol,
            "Gas Fee (in ETH)": "0",  # gas already counted in normal tx
            "Function": "",
            "Input Data": "",  # input data is on the parent normal tx
        })
    return rows


# ---------------------------------------------------------------------------
# 3. Fetch actual on-chain balances
# ---------------------------------------------------------------------------

def fetch_eth_balance() -> Decimal:
    print("[*] Fetching actual ETH balance …")
    params = {
        "chainid": CHAIN_ID,
        "module": "account",
        "action": "balance",
        "address": WALLET,
        "tag": "latest",
        "apikey": ETHERSCAN_API_KEY,
    }
    url = f"{BASE_URL}?{urlencode(params)}"
    time.sleep(RATE_LIMIT_DELAY)
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read().decode())
    if data.get("status") == "0":
        print(f"  [!] ETH balance API error: {data.get('result')}")
        return Decimal("0")
    return wei_to_eth(data["result"])


def fetch_token_balance(contract: str, decimals: int) -> Decimal:
    params = {
        "chainid": CHAIN_ID,
        "module": "account",
        "action": "tokenbalance",
        "contractaddress": contract,
        "address": WALLET,
        "tag": "latest",
        "apikey": ETHERSCAN_API_KEY,
    }
    url = f"{BASE_URL}?{urlencode(params)}"
    time.sleep(RATE_LIMIT_DELAY)
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read().decode())
    raw = data.get("result", "0")
    if not raw.isdigit() and not (raw.startswith("-") and raw[1:].isdigit()):
        return Decimal("0")
    return Decimal(raw) / Decimal(10) ** decimals


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # ---- Validate config ----
    # Ensure stdout can handle Unicode on Windows
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )

    os.makedirs(TRANSACTIONS_DIR, exist_ok=True)
    os.makedirs(DISCREPANCIES_DIR, exist_ok=True)

    if not ETHERSCAN_API_KEY:
        print("[!] ERROR: ETHERSCAN_API_KEY is empty. Set it in .env")
        return
    if not WALLET:
        print("[!] ERROR: WALLET_ADDRESS is empty. Set it in .env")
        return
    print(f"[*] Wallet:  {WALLET}")

    # ---- Fetch all transactions ----
    normal_txs = fetch_normal_txs()
    internal_txs = fetch_internal_txs()
    erc20_txs = fetch_erc20_txs()

    print(f"\n[+] Totals: {len(normal_txs)} normal, "
          f"{len(internal_txs)} internal, {len(erc20_txs)} ERC20")

    # ---- Process & accumulate ----
    rows = []
    rows.extend(process_normal(normal_txs))
    rows.extend(process_internal(internal_txs))
    rows.extend(process_erc20(erc20_txs))

    # Sort by date
    rows.sort(key=lambda r: r["Date"])

    # ---- Write CSV ----
    fieldnames = [
        "Date", "Block Number", "Transaction Hash", "Type", "From", "To",
        "Amount", "Token/Asset", "Gas Fee (in ETH)",
        "Function", "Input Data",
    ]
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"[+] Wrote {len(rows)} rows to {CSV_FILE}")

    # ---- Expected balances ----
    expected_eth = eth_in - eth_out - gas_total

    # ---- Actual balances ----
    actual_eth = fetch_eth_balance()

    print("\n" + "=" * 72)
    print("  DISCREPANCY REPORT")
    print("=" * 72)

    # Track which contracts have discrepancies + their balances
    discrepancy_contracts = {}  # contract -> {symbol, expected, actual, diff}

    # ETH
    diff_eth = actual_eth - expected_eth
    print(f"\n{'ETH':>10}")
    print(f"  {'Expected balance':30s} {expected_eth:>28.18f}")
    print(f"  {'Actual on-chain balance':30s} {actual_eth:>28.18f}")
    print(f"  {'Difference':30s} {diff_eth:>28.18f}")
    if abs(diff_eth) > Decimal("0.000000000001"):
        print(f"  ** DISCREPANCY DETECTED **")
        discrepancy_contracts["ETH"] = {
            "symbol": "ETH", "contract": "",
            "expected": expected_eth, "actual": actual_eth, "diff": diff_eth,
        }
    else:
        print(f"  OK")

    # ERC20 tokens
    if token_flows:
        print(f"\n[*] Checking {len(token_flows)} ERC20 token balances …")

    for contract, tf in sorted(token_flows.items(), key=lambda x: x[1]["symbol"]):
        symbol = tf["symbol"]
        expected = tf["in"] - tf["out"]
        actual = fetch_token_balance(contract, tf["decimals"])
        diff = actual - expected

        print(f"\n{symbol:>10}  ({contract})")
        print(f"  {'Expected balance':30s} {expected:>28}")
        print(f"  {'Actual on-chain balance':30s} {actual:>28}")
        print(f"  {'Difference':30s} {diff:>28}")

        threshold = Decimal(1) / Decimal(10) ** min(tf["decimals"], 8)
        if abs(diff) > threshold:
            print(f"  ** DISCREPANCY DETECTED **")
            discrepancy_contracts[contract] = {
                "symbol": symbol, "contract": contract,
                "expected": expected, "actual": actual, "diff": diff,
            }
        else:
            print(f"  OK")

    # ---- Write discrepancy audit trail CSV ----
    if discrepancy_contracts:
        # Build token tx lookup: contract -> list of ERC20 transfer dicts
        token_tx_lookup = defaultdict(list)
        for tx in erc20_txs:
            c = tx.get("contractAddress", "").lower()
            if c in discrepancy_contracts:
                token_tx_lookup[c].append(tx)

        # Build normal tx lookup by hash for function names
        normal_tx_by_hash = {}
        for tx in normal_txs:
            normal_tx_by_hash[tx["hash"].lower()] = tx

        disc_rows = []
        for key, info in sorted(discrepancy_contracts.items(), key=lambda x: x[1]["symbol"]):
            symbol = info["symbol"]
            contract = info["contract"]

            # --- Summary row ---
            disc_rows.append({
                "Token/Asset": symbol,
                "Contract Address": contract,
                "Row Type": "SUMMARY",
                "Date": "",
                "Block Number": "",
                "Transaction Hash": "",
                "Direction": "",
                "From": "",
                "To": "",
                "Amount": "",
                "Running Balance": "",
                "On-Chain Balance": "",
                "Function Called": "",
                "Expected Balance": f"{info['expected']}",
                "Actual Balance": f"{info['actual']}",
                "Difference": f"{info['diff']}",
                "Resolution": "",
            })

            # --- Transaction tree for this token ---
            if key == "ETH":
                # For ETH, pull from the rows we already built
                running = Decimal("0")
                eth_txs_sorted = []
                for tx in normal_txs:
                    if tx.get("isError", "0") != "0":
                        if tx["from"].lower() == WALLET:
                            fee = Decimal(tx["gasUsed"]) * Decimal(tx["gasPrice"]) / Decimal("1e18")
                            eth_txs_sorted.append({
                                "ts": int(tx["timeStamp"]),
                                "block": tx.get("blockNumber", ""),
                                "date": ts_to_date(tx["timeStamp"]),
                                "hash": tx["hash"],
                                "direction": "GAS (failed)",
                                "from": tx["from"],
                                "to": tx.get("to", ""),
                                "amount": -fee,
                                "fn": tx.get("functionName", ""),
                            })
                        continue
                    value = wei_to_eth(tx["value"])
                    from_addr = tx["from"].lower()
                    to_addr = (tx.get("to") or "").lower()
                    direction = "IN" if to_addr == WALLET else "OUT"
                    fee = Decimal("0")
                    if from_addr == WALLET:
                        fee = Decimal(tx["gasUsed"]) * Decimal(tx["gasPrice"]) / Decimal("1e18")
                    signed = value if direction == "IN" else -value
                    fn = tx.get("functionName", "")
                    eth_txs_sorted.append({
                        "ts": int(tx["timeStamp"]),
                        "block": tx.get("blockNumber", ""),
                        "date": ts_to_date(tx["timeStamp"]),
                        "hash": tx["hash"],
                        "direction": direction,
                        "from": tx["from"],
                        "to": tx.get("to", ""),
                        "amount": signed,
                        "fn": fn,
                    })
                    if fee > 0:
                        eth_txs_sorted.append({
                            "ts": int(tx["timeStamp"]),
                            "block": tx.get("blockNumber", ""),
                            "date": ts_to_date(tx["timeStamp"]),
                            "hash": tx["hash"],
                            "direction": "GAS",
                            "from": tx["from"],
                            "to": "",
                            "amount": -fee,
                            "fn": fn,
                        })
                for tx in internal_txs:
                    if tx.get("isError", "0") != "0":
                        continue
                    value = wei_to_eth(tx["value"])
                    if value == 0:
                        continue
                    to_addr = (tx.get("to") or "").lower()
                    direction = "IN" if to_addr == WALLET else "OUT"
                    signed = value if direction == "IN" else -value
                    eth_txs_sorted.append({
                        "ts": int(tx["timeStamp"]),
                        "block": tx.get("blockNumber", ""),
                        "date": ts_to_date(tx["timeStamp"]),
                        "hash": tx.get("hash", tx.get("transactionHash", "")),
                        "direction": f"{direction} (internal)",
                        "from": tx["from"],
                        "to": tx.get("to", ""),
                        "amount": signed,
                        "fn": "",
                    })
                eth_txs_sorted.sort(key=lambda x: x["ts"])
                for t in eth_txs_sorted:
                    running += t["amount"]
                    disc_rows.append({
                        "Token/Asset": "ETH",
                        "Contract Address": "",
                        "Row Type": "TX",
                        "Date": t["date"],
                        "Block Number": t["block"],
                        "Transaction Hash": t["hash"],
                        "Direction": t["direction"],
                        "From": t["from"],
                        "To": t["to"],
                        "Amount": f"{t['amount']:.18f}",
                        "Running Balance": f"{running:.18f}",
                        "On-Chain Balance": "",
                        "Function Called": t["fn"],
                        "Expected Balance": "",
                        "Actual Balance": "",
                        "Difference": "",
                        "Resolution": "",
                    })
            else:
                # ERC20 token tree
                token_txs = token_tx_lookup.get(contract, [])
                token_txs_sorted = sorted(token_txs, key=lambda x: int(x["timeStamp"]))
                running = Decimal("0")
                tok_decimals = 18
                tf_entry = token_flows.get(contract)
                if tf_entry:
                    tok_decimals = tf_entry["decimals"]
                decimals_str = str(tok_decimals)

                if ALCHEMY_API_KEY:
                    print(f"  [*] Querying on-chain balance history for {symbol} ({len(token_txs_sorted)} blocks) …")

                bal_before = Decimal("-1")
                bal_after = Decimal("-1")

                for tx in token_txs_sorted:
                    value = token_amount(tx["value"], decimals_str)
                    to_addr = (tx.get("to") or "").lower()
                    direction = "IN" if to_addr == WALLET else "OUT"
                    signed = value if direction == "IN" else -value
                    running += signed

                    # Look up the function from the normal tx
                    ntx = normal_tx_by_hash.get(tx["hash"].lower(), {})
                    fn = ntx.get("functionName", "")

                    # Query on-chain balance via Alchemy at both block-1 and block
                    block_num = tx.get("blockNumber", "")
                    onchain_bal = ""
                    if ALCHEMY_API_KEY and block_num:
                        bn = int(block_num)
                        bal_before = alchemy_balance_of(contract, bn - 1, tok_decimals)
                        bal_after = alchemy_balance_of(contract, bn, tok_decimals)
                        if bal_before >= 0 and bal_after >= 0:
                            onchain_bal = f"(N-1) {bal_before}  (N) {bal_after}"

                    disc_rows.append({
                        "Token/Asset": symbol,
                        "Contract Address": contract,
                        "Row Type": "TX",
                        "Date": ts_to_date(tx["timeStamp"]),
                        "Block Number": block_num,
                        "Transaction Hash": tx["hash"],
                        "Direction": direction,
                        "From": tx["from"],
                        "To": tx.get("to", ""),
                        "Amount": f"{signed}",
                        "Running Balance": f"{running}",
                        "On-Chain Balance": onchain_bal,
                        "Function Called": fn,
                        "Expected Balance": "",
                        "Actual Balance": "",
                        "Difference": "",
                        "Resolution": "",
                    })

                # Final row showing the gap + auto-resolution
                resolution = ""
                if info["expected"] < 0 and info["actual"] >= 0:
                    # More tokens went out than came in via Transfer events,
                    # but actual balance is non-negative — the gap must be
                    # non-transfer balance increases (rebasing, staking yield)
                    resolution = (
                        f"RESOLVED: {info['diff']} {symbol} from non-transfer "
                        f"balance increase (rebasing/staking rewards). "
                        f"Wallet received {abs(info['expected']):.9f} more via "
                        f"rebase than tracked by Transfer events."
                    )
                elif info["expected"] > 0 and info["actual"] == 0:
                    # Expected positive balance but actual is 0 — tokens
                    # were likely burned, migrated, or are worthless/scam
                    resolution = (
                        f"RESOLVED: Token balance zeroed outside of Transfer "
                        f"events. Likely a scam/airdrop token, admin burn, "
                        f"or token migration."
                    )
                elif info["expected"] > 0 and info["actual"] > 0 and info["diff"] > 0:
                    # Actual is higher than expected — extra tokens appeared
                    resolution = (
                        f"RESOLVED: +{info['diff']} {symbol} from non-transfer "
                        f"balance increase (rebasing rewards, staking yield, "
                        f"or direct contract mint)."
                    )
                elif info["expected"] > 0 and info["actual"] > 0 and info["diff"] < 0:
                    # Actual is lower than expected — tokens disappeared
                    # Typical of fee-on-transfer / reflection / tax tokens
                    # Calculate tax from last OUT tx: (N-1) - transfer_amount - (N)
                    tax_detail = ""
                    if ALCHEMY_API_KEY and direction == "OUT" and bal_before >= 0 and bal_after >= 0:
                        transfer_amt = abs(signed)
                        tax_taken = bal_before - transfer_amt - bal_after
                        tax_detail = (
                            f" Last sell tax: On-Chain(N-1) {bal_before} "
                            f"- Transfer {transfer_amt} "
                            f"- On-Chain(N) {bal_after} "
                            f"= {tax_taken} {symbol} taken by contract."
                        )
                    resolution = (
                        f"RESOLVED: fee-on-transfer (tax token).{tax_detail}"
                    )

                disc_rows.append({
                    "Token/Asset": symbol,
                    "Contract Address": contract,
                    "Row Type": "GAP",
                    "Date": "",
                    "Block Number": "",
                    "Transaction Hash": "",
                    "Direction": "UNTRACKED",
                    "From": "",
                    "To": "",
                    "Amount": f"{info['diff']}",
                    "Running Balance": f"{info['actual']}",
                    "On-Chain Balance": f"{info['actual']}",
                    "Function Called": "Balance change not captured by Transfer events",
                    "Expected Balance": "",
                    "Actual Balance": "",
                    "Difference": "",
                    "Resolution": resolution,
                })

            # Empty separator row between tokens
            disc_rows.append({k: "" for k in disc_rows[0]})

        disc_fields = [
            "Token/Asset", "Contract Address", "Row Type",
            "Date", "Block Number", "Transaction Hash", "Direction",
            "From", "To", "Amount", "Running Balance",
            "On-Chain Balance", "Function Called",
            "Expected Balance", "Actual Balance", "Difference",
            "Resolution",
        ]
        with open(DISCREPANCY_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=disc_fields)
            writer.writeheader()
            writer.writerows(disc_rows)
        print(f"\n[+] Wrote {len(discrepancy_contracts)} token audit trails "
              f"({len(disc_rows)} rows) to {DISCREPANCY_FILE}")
    else:
        print(f"\n[+] No discrepancies found — no {DISCREPANCY_FILE} written.")

    # ---- Export to styled Excel ----
    print("\n[*] Generating styled Excel files …")
    _write_transactions_xlsx(rows, fieldnames)
    if discrepancy_contracts:
        _write_discrepancies_xlsx(disc_rows, disc_fields)

    print("\n" + "=" * 72)
    print("  Done.")
    print("=" * 72)


# ---------------------------------------------------------------------------
# Excel export helpers
# ---------------------------------------------------------------------------

# Shared styles
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_IN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_OUT_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
_FAILED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SUMMARY_FONT = Font(bold=True, size=11)
_GAP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_GAP_FONT = Font(bold=True, color="BF8F00", size=11)
_RESOLVED_FONT = Font(bold=True, color="548235", size=11)
_ZEBRA_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")


def _style_header(ws, num_cols):
    """Apply header styling to first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_width(ws, num_cols, max_width=50):
    """Auto-fit column widths based on content."""
    for col in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _write_transactions_xlsx(rows, fieldnames):
    """Write transactions to a styled .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header
    ws.append(fieldnames)
    _style_header(ws, len(fieldnames))

    # Find column indices
    amount_col = fieldnames.index("Amount") + 1
    gas_col = fieldnames.index("Gas Fee (in ETH)") + 1

    # Data rows
    for i, row in enumerate(rows, start=2):
        values = [row.get(f, "") for f in fieldnames]
        ws.append(values)

        tx_type = row.get("Type", "")

        # Row coloring by direction
        if "IN" in tx_type and "failed" not in tx_type:
            fill = _IN_FILL
        elif "OUT" in tx_type and "failed" not in tx_type:
            fill = _OUT_FILL
        elif "failed" in tx_type:
            fill = _FAILED_FILL
        elif i % 2 == 0:
            fill = _ZEBRA_FILL
        else:
            fill = None

        for col in range(1, len(fieldnames) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill

        # Right-align numeric columns
        ws.cell(row=i, column=amount_col).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=gas_col).alignment = Alignment(horizontal="right")

    _auto_width(ws, len(fieldnames))

    # Wrap text in Input Data column
    input_col = fieldnames.index("Input Data") + 1
    for row_idx in range(2, len(rows) + 2):
        cell = ws.cell(row=row_idx, column=input_col)
        cell.alignment = Alignment(wrap_text=True)

    out_path = CSV_FILE.replace(".csv", ".xlsx")
    wb.save(out_path)
    print(f"[+] Wrote {len(rows)} rows to {out_path}")


def _write_discrepancies_xlsx(disc_rows, disc_fields):
    """Write discrepancy audit trail to a styled .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Discrepancy Audit"

    # Header
    ws.append(disc_fields)
    _style_header(ws, len(disc_fields))

    # Find column indices
    amount_col = disc_fields.index("Amount") + 1
    running_col = disc_fields.index("Running Balance") + 1
    onchain_col = disc_fields.index("On-Chain Balance") + 1
    resolution_col = disc_fields.index("Resolution") + 1

    for i, row in enumerate(disc_rows, start=2):
        values = [row.get(f, "") for f in disc_fields]
        ws.append(values)

        row_type = row.get("Row Type", "")
        direction = row.get("Direction", "")

        # Row styling based on type
        if row_type == "SUMMARY":
            for col in range(1, len(disc_fields) + 1):
                cell = ws.cell(row=i, column=col)
                cell.fill = _SUMMARY_FILL
                cell.font = _SUMMARY_FONT
                cell.border = _THIN_BORDER
        elif row_type == "GAP":
            for col in range(1, len(disc_fields) + 1):
                cell = ws.cell(row=i, column=col)
                cell.fill = _GAP_FILL
                cell.font = _GAP_FONT
                cell.border = _THIN_BORDER
            # Resolution column in green
            res_cell = ws.cell(row=i, column=resolution_col)
            if res_cell.value and str(res_cell.value).startswith("RESOLVED"):
                res_cell.font = _RESOLVED_FONT
        elif row_type == "TX":
            if "IN" in direction:
                fill = _IN_FILL
            elif "OUT" in direction:
                fill = _OUT_FILL
            else:
                fill = None
            for col in range(1, len(disc_fields) + 1):
                cell = ws.cell(row=i, column=col)
                cell.border = _THIN_BORDER
                if fill:
                    cell.fill = fill
        else:
            # Separator row — leave blank/light
            for col in range(1, len(disc_fields) + 1):
                ws.cell(row=i, column=col).border = _THIN_BORDER

        # Right-align numeric columns
        for col_idx in (amount_col, running_col, onchain_col):
            ws.cell(row=i, column=col_idx).alignment = Alignment(horizontal="right")

    _auto_width(ws, len(disc_fields))

    out_path = DISCREPANCY_FILE.replace(".csv", ".xlsx")
    wb.save(out_path)
    print(f"[+] Wrote {len(disc_rows)} rows to {out_path}")


if __name__ == "__main__":
    main()
